import cx_Oracle
import pandas as pd
import csv
import datetime as dt
import win32com.client
import xlsxwriter
from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import os

dsn_tns = cx_Oracle.makedsn('psdb33540-vip.lexis-nexis.com', '1521', service_name='SBL_SERVICE.isprod.lexisnexis.com') 
conn = cx_Oracle.connect(user='SDI', password='SDIPROD63', dsn=dsn_tns) 

query_agr = """-- Agreement Level
SELECT
  ORG.LOC AS "Customer"
  , ORG.NAME AS "Customer Name"
  , ORG.X_CUSTOMER_SUBCLASS AS "Subclass", ORG.X_SEC_SUBCLASS AS "Secondary Subclass"
  , EMP.LAST_NAME AS "Last Name", EMP.FST_NAME AS "First Name"
  , AGR.AGREE_NUM||':'||AGR.REV_NUM AS "Agreement", AGR.X_MULTI_TERM_FLG AS MYD, AGR.X_MYD_MAX_END_DT AS "MYD End"
  , AGR.EFF_START_DT AS "Sub Start", AGR.EFF_END_DT AS "Sub End"
  , AGR_CON.LAST_NAME AS "Contact Last Name", AGR_CON.FST_NAME AS "Contact First Name", AGR_CON.EMAIL_ADDR AS "Contact Email", AGR_CON.WORK_PH_NUM AS "Contact Phone"
  , ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1) AS "Length"
  , AGR.X_TERM_TYPE AS "Term"
  , AGR.X_AAR_PCT AS "AAR%"
  , ORD.X_ORDER_SUB_TYPE AS "Order Type"
  , AGR.BL_CURCY_CD AS FX, AGR.X_MONTLY_NET_PRICE AS "Total Agreement Value"
  ,round(AGR.X_MONTLY_NET_PRICE / ( ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1) ),2) AS "Monthly Commitment"
  , ROUND(SUM(NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI)), 2) AS "Value (Doc)"
  , ROUND(SUM(NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI)) * (12/NULLIF(ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1), 0)), 2) AS "Annualized (Doc)"
  , ROUND(SUM(NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI)), 2) AS "Renewal (Doc)"
  , ROUND(SUM(NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI)) * (12/NULLIF(ROUND(MONTHS_BETWEEN(REL_Q.EFF_END_DT+1, REL_Q.EFF_START_DT), 1), 0)), 2) AS "Annualized Renewal (Doc)"
  , ROUND(SUM(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI)), 2) AS "Value (USD)"
  , ROUND(SUM(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI)) * (12/NULLIF(ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1), 0)), 2) AS "Annualized (USD)"
  , ROUND(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * SUM(NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI)), 2) AS "Renewal (USD)"
  , ROUND(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * SUM(NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI)) * (12/NULLIF(ROUND(MONTHS_BETWEEN(REL_Q.EFF_END_DT+1, REL_Q.EFF_START_DT), 1), 0)), 2) AS "Annualized Renewal (USD)"
FROM
  SIEBEL.S_DOC_AGREE AGR
  INNER JOIN SIEBEL.S_ORG_EXT ORG ON ORG.ROW_ID = AGR.TARGET_OU_ID
  INNER JOIN SIEBEL.S_BU BU ON BU.ROW_ID = ORG.BU_ID
  INNER JOIN SIEBEL.S_ORDER ORD ON ORD.ROW_ID = AGR.ORDER_ID
  INNER JOIN SIEBEL.S_ORDER_ITEM OI ON OI.ORDER_ID = ORD.ROW_ID
  INNER JOIN SIEBEL.S_PROD_INT PROD ON PROD.ROW_ID = OI.PROD_ID
  INNER JOIN SIEBEL.S_POSTN POS ON POS.ROW_ID = ORG.PR_POSTN_ID
  INNER JOIN SIEBEL.S_CONTACT EMP ON EMP.ROW_ID = POS.PR_EMP_ID
  INNER JOIN SIEBEL.S_CONTACT AGR_CON ON AGR_CON.PAR_ROW_ID = AGR.CON_PER_ID
  LEFT OUTER JOIN SIEBEL.S_DOC_QUOTE REL_Q ON REL_Q.X_REL_QUOTE_ID = AGR.QUOTE_ID
  LEFT OUTER JOIN SIEBEL.S_QUOTE_ITEM REL_QI ON REL_QI.SD_ID = REL_Q.ROW_ID AND REL_QI.PROD_ID = OI.PROD_ID
WHERE 1=1
  AND BU.NAME = 'United Kingdom'
  AND NVL(AGR.X_TRIAL_FLG, 'N') = 'N'     -- No Trials
  AND AGR.EFF_END_DT > SYSDATE            -- Renews in future
  AND AGR.STAT_CD = 'Active'              -- Active
  AND PROD.INTEGRATION_ID NOT IN
    (
      'urn:product:1523691','urn:product:1523692','urn:product:1523693','urn:product:1523694','urn:product:1523695','urn:product:1523696','urn:product:1523697', 'urn:product:1523698'
      ,'urn:product:1523699','urn:product:1523700','urn:product:1525008','urn:product:1525009','urn:product:1525010','urn:product:1525618','urn:product:1526114' -- Remove international content
      ,'urn:product:1515564' -- Remove UK Core Feature
    )
GROUP BY
  ORG.LOC
  , ORG.NAME
  , ORG.X_CUSTOMER_SUBCLASS, ORG.X_SEC_SUBCLASS
  , EMP.LAST_NAME, EMP.FST_NAME
  , AGR.AGREE_NUM
  , AGR.REV_NUM
  , AGR.X_MULTI_TERM_FLG, AGR.X_MYD_MAX_END_DT
  , AGR.EFF_START_DT, AGR.EFF_END_DT
  , REL_Q.EFF_START_DT, REL_Q.EFF_END_DT
  , AGR_CON.LAST_NAME, AGR_CON.FST_NAME, AGR_CON.EMAIL_ADDR, AGR_CON.WORK_PH_NUM
  , AGR.X_TERM_TYPE
  , AGR.X_AAR_PCT
  , ORD.X_ORDER_SUB_TYPE
  , AGR.BL_CURCY_CD, AGR.X_MONTLY_NET_PRICE
ORDER BY
  ORG.LOC, AGR.EFF_END_DT, AGR.EFF_START_DT, AGR.AGREE_NUM, AGR.REV_NUM"""


query_prod = """SELECT
  ORG.LOC AS "Customer"
  , ORG.NAME AS "Customer Name"
  , ORG.X_CUSTOMER_SUBCLASS AS "Subclass", ORG.X_SEC_SUBCLASS AS "Secondary Subclass"
  , EMP.LAST_NAME AS "Last Name", EMP.FST_NAME AS "First Name", EMP.job_title as "Position"
  , AGR.AGREE_NUM||':'||AGR.REV_NUM AS "Agreement", AGR.X_MULTI_TERM_FLG AS MYD, AGR.X_MYD_MAX_END_DT AS "MYD End"
  , AGR.EFF_START_DT AS "Sub Start", AGR.EFF_END_DT AS "Sub End"
  , AGR_CON.LAST_NAME AS "Contact Last Name", AGR_CON.FST_NAME AS "Contact First Name", AGR_CON.EMAIL_ADDR AS "Contact Email", AGR_CON.WORK_PH_NUM AS "Contact Phone"
  , ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1) AS "Length"
  , AGR.X_TERM_TYPE AS "Term"
  , AGR.X_AAR_PCT AS "AAR%"
  , ORD.X_ORDER_SUB_TYPE AS "Order Type"
  --, MAX(CASE WHEN PROD.INTEGRATION_ID = 'urn:product:1519442' THEN 'Y' ELSE 'N' END) OVER (PARTITION BY ORG.ROW_ID, AGR.ROW_ID) AS "Has Nexis Core Feature"
  , PROD.INTEGRATION_ID AS "Product PGUID", PROD.NAME AS "Product Name", PROD.X_PRODUCT_CLASS AS "Product Class", PROD.X_GO_TO_MARKET_TYPE AS "GTM Type", PROD.X_ITEM_TYPE AS "Item Type", PROD.X_PKG_LVL AS "Package Level", PROD.X_PRODUCT_TYPE AS "Product Type"
  , AGR.BL_CURCY_CD AS FX, AGR.X_MONTLY_NET_PRICE AS "Total Agreement Value", round(AGR.X_MONTLY_NET_PRICE /  (ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1) ),2) as "Monthly Commitment"
  , NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI) AS "Value (Doc)"
  , ROUND(NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI) * (12/NULLIF(ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1), 0)), 2) AS "Annualized (Doc)"
  , NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI) AS "Renewal (Doc)"
  , ROUND(NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI) * (12/NULLIF(ROUND(MONTHS_BETWEEN(REL_Q.EFF_END_DT+1, REL_Q.EFF_START_DT), 1), 0)), 2) AS "Annualized Renewal (Doc)"
  , ROUND(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI), 2) AS "Value (USD)"
  , ROUND(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * NVL(OI.X_EXT_NET_PRI, OI.QTY_REQ * OI.NET_PRI) * (12/NULLIF(ROUND(MONTHS_BETWEEN(AGR.EFF_END_DT+1, AGR.EFF_START_DT), 1), 0)), 2) AS "Annualized (USD)"
  , ROUND(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI), 2) AS "Renewal (USD)"
  , ROUND(CASE AGR.BL_CURCY_CD WHEN 'GBP' THEN 1.28 WHEN 'EUR' THEN 1.12281 ELSE 1 END * NVL(REL_QI.X_EXT_NET_PRI, REL_QI.QTY_REQ * REL_QI.NET_PRI) * (12/NULLIF(ROUND(MONTHS_BETWEEN(REL_Q.EFF_END_DT+1, REL_Q.EFF_START_DT), 1), 0)), 2) AS "Annualized Renewal (USD)"
FROM
  SIEBEL.S_DOC_AGREE AGR
  INNER JOIN SIEBEL.S_ORG_EXT ORG ON ORG.ROW_ID = AGR.TARGET_OU_ID
  INNER JOIN SIEBEL.S_BU BU ON BU.ROW_ID = ORG.BU_ID
  INNER JOIN SIEBEL.S_ORDER ORD ON ORD.ROW_ID = AGR.ORDER_ID
  INNER JOIN SIEBEL.S_ORDER_ITEM OI ON OI.ORDER_ID = ORD.ROW_ID
  INNER JOIN SIEBEL.S_PROD_INT PROD ON PROD.ROW_ID = OI.PROD_ID
  INNER JOIN SIEBEL.S_POSTN POS ON POS.ROW_ID = ORG.PR_POSTN_ID
  INNER JOIN SIEBEL.S_CONTACT EMP ON EMP.ROW_ID = POS.PR_EMP_ID
  INNER JOIN SIEBEL.S_CONTACT AGR_CON ON AGR_CON.PAR_ROW_ID = AGR.CON_PER_ID
  LEFT OUTER JOIN SIEBEL.S_DOC_QUOTE REL_Q ON REL_Q.X_REL_QUOTE_ID = AGR.QUOTE_ID
  LEFT OUTER JOIN SIEBEL.S_QUOTE_ITEM REL_QI ON REL_QI.SD_ID = REL_Q.ROW_ID AND REL_QI.PROD_ID = OI.PROD_ID
WHERE 1=1
  AND BU.NAME = 'United Kingdom'
  AND NVL(AGR.X_TRIAL_FLG, 'N') = 'N'     -- No Trials
  AND AGR.EFF_END_DT > SYSDATE            -- Renews in future
  AND AGR.STAT_CD = 'Active'              -- Active
  AND PROD.INTEGRATION_ID NOT IN
    (
      'urn:product:1523691','urn:product:1523692','urn:product:1523693','urn:product:1523694','urn:product:1523695','urn:product:1523696','urn:product:1523697', 'urn:product:1523698'
      ,'urn:product:1523699','urn:product:1523700','urn:product:1525008','urn:product:1525009','urn:product:1525010','urn:product:1525618','urn:product:1526114' -- Remove international content
      ,'urn:product:1515564' -- Remove UK Core Feature
    )
ORDER BY
  ORG.LOC, AGR.EFF_END_DT, AGR.EFF_START_DT, AGR.AGREE_NUM, AGR.REV_NUM, PROD.INTEGRATION_ID"""



agr = pd.read_sql(query_agr, con=conn)

prod = pd.read_sql(query_prod, con=conn)

mapping = pd.read_csv('Mapping.csv')

print('Fetching data...')

today = dt.datetime.today().strftime('%Y%m%d')


#output_file_agr = 'Agreement Level {}.csv'.format(today)
#agr.to_csv(output_file_agr, index=False)

#output_file_prod = 'Product Level {}.csv'.format(today)
#prod.to_csv(output_file_prod, index=False)



wb = Workbook()
prod_sheet = wb.active
prod_sheet.title = 'Product Level'
agr_sheet = wb.create_sheet('Agreement Level')
mapping_sheet = wb.create_sheet('Mapping')

# PRODUCT LEVEL
for r in dataframe_to_rows(prod, index=False):
    try:
        prod_sheet.append(r)
    except KeyError:
        prod_sheet.append(r[:9]+['']+r[10:])
#dates as dates
from openpyxl.styles import NamedStyle
date_style = NamedStyle(name='date_style',number_format='dd-mmm-yy')
wb.add_named_style(date_style)
for row in prod_sheet.iter_rows(min_col=9, max_col=12):
    for cell in row:
        cell.style = date_style
# format numbers
for row in prod_sheet.iter_rows(min_col=29, max_col=38):
    for cell in row:
        cell.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'

# expand columns
ws = prod_sheet
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
for col, value in dims.items():
    ws.column_dimensions[col].width = value+5
#bold headings and freeze panes
for cell in prod_sheet[1]:
    cell.font = Font(bold=True)
prod_sheet.freeze_panes = "A2"


#AGREEMENT LEVEL
for r in dataframe_to_rows(agr, index=False):
    try:
        agr_sheet.append(r)
    except KeyError:
        agr_sheet.append(r[:8]+['']+r[9:])
#dates as dates
for row in agr_sheet.iter_rows(min_col=9, max_col=12):
    for cell in row:
        cell.style = date_style
# format numbers
for row in agr_sheet.iter_rows(min_col=21, max_col=30):
    for cell in row:
        cell.number_format = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
# expand columns
ws = agr_sheet
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
for col, value in dims.items():
    ws.column_dimensions[col].width = value+5
#bold headings and freeze panes
for cell in agr_sheet[1]:
    cell.font = Font(bold=True)
agr_sheet.freeze_panes = "A2"



#  MAPPING
for r in dataframe_to_rows(mapping, index=False):
    mapping_sheet.append(r)
for cell in mapping_sheet[1]:
    cell.font = Font(bold=True)


# ZOOM TO 85%

for ws in wb.worksheets:
    ws.sheet_view.zoomScale = 85



wb.save('{} - COMPANY CONFIDENTIAL - UK GCRM Agreements.xlsx'.format(today))



for f in os.listdir():
  if f.endswith('xlsx'):
    UKGCRM = os.path.dirname(os.path.realpath(__file__))+"\\" +f


import win32com.client as win32

excel = win32.gencache.EnsureDispatch('Excel.Application')
#Before saving the file set DisplayAlerts to False to suppress the warning dialog:
excel.DisplayAlerts = False
wb = excel.Workbooks.Open(UKGCRM)
# refer https://docs.microsoft.com/en-us/previous-versions/office/developer/office-2007/bb214129(v=office.12)?redirectedfrom=MSDN
# FileFormat = 51 is for .xlsx extension
wb.SaveAs(UKGCRM, 51, 'UK_G$CRM')                                               
wb.Close() 
excel.Application.Quit()
